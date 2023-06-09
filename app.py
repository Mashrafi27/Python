import openpyxl as xl
wb = xl.load_workbook("Python Tutorial Supplementary Materials/transactions.xlsx")
sheet = wb['Sheet1']
print(sheet['a1'].value)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    print(cell.value)
    corrected_price = cell.value*0.9
    corrected_price_cell = sheet.cell(row,4)
    corrected_price_cell.value = corrected_price

wb.save('transactions2.xlsx')